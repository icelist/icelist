"""数据落地：JSON / Excel（双语表头 + 嵌入图片 + 多 Sheet 分类）."""
from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Iterable

from loguru import logger
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from tqdm import tqdm

from .base import Product
from .i18n import (
    PLATFORM_RU, TYPE_RU, bilingual, price_bucket_ru, specs_bilingual,
)
from .utils import download_image, safe_filename

try:
    from PIL import Image as PILImage
    HAS_PIL = True
except Exception:
    HAS_PIL = False


THUMB_SIZE = 90
EXCEL_ROW_HEIGHT = 70
PREVIEW_COL_WIDTH = 14

HEADER_FILL = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFFFF", bold=True, size=10)


PREVIEW_KEY = "_preview"

COLUMNS: list[tuple[str, str, object]] = [
    (PREVIEW_KEY,      bilingual("图片预览"),  None),
    ("platform_zh",    bilingual("平台"),      lambda p: p.platform),
    ("platform_ru",    "Platform RU",          lambda p: PLATFORM_RU.get(p.platform, p.platform)),
    ("product_id",     bilingual("商品ID"),    lambda p: p.product_id),
    ("title",          bilingual("标题"),      lambda p: p.title or ""),
    ("price",          bilingual("价格"),      lambda p: p.price if p.price is not None else ""),
    ("price_text",     bilingual("价格文案"),  lambda p: p.price_text or ""),
    ("bucket_type_zh", bilingual("类型"),      lambda p: p.bucket_type or ""),
    ("bucket_type_ru", bilingual("类型(俄)"),  lambda p: TYPE_RU.get(p.bucket_type or "", "")),
    ("bucket_price",   bilingual("价格区间"),  lambda p: f"{p.bucket_price or ''} / {price_bucket_ru(p.bucket_price)}" if p.bucket_price else ""),
    ("category_path",  bilingual("平台类目"),  lambda p: p.category_path or ""),
    ("shop",           bilingual("店铺"),      lambda p: p.shop or ""),
    ("sales",          bilingual("销量"),      lambda p: p.sales or ""),
    ("brand",          bilingual("品牌"),      lambda p: p.brand or ""),
    ("origin",         bilingual("产地"),      lambda p: p.origin or ""),
    ("material",       bilingual("材质"),      lambda p: p.material or ""),
    ("moq",            bilingual("起订量"),    lambda p: p.moq or ""),
    ("delivery",       bilingual("发货"),      lambda p: p.delivery or ""),
    ("keyword",        bilingual("搜索词"),    lambda p: p.keyword or ""),
    ("features",       bilingual("卖点/特点"), lambda p: " | ".join(p.features)),
    ("specs",          bilingual("规格"),      lambda p: specs_bilingual(p.specs)),
    ("description",    bilingual("描述"),      lambda p: (p.description or "")[:500]),
    ("images",         bilingual("图片URL"),   lambda p: " | ".join(p.images)),
    ("local_images",   bilingual("本地图片"),  lambda p: " | ".join(p.local_images)),
    ("url",            bilingual("详情链接"),  lambda p: p.url),
]

COL_WIDTHS = {
    PREVIEW_KEY: PREVIEW_COL_WIDTH,
    "platform_zh": 12, "platform_ru": 18, "product_id": 16,
    "title": 50, "price": 10, "price_text": 14,
    "bucket_type_zh": 12, "bucket_type_ru": 16, "bucket_price": 22,
    "category_path": 24, "shop": 20, "sales": 12,
    "brand": 14, "origin": 14, "material": 14,
    "moq": 14, "delivery": 14, "keyword": 14,
    "features": 40, "specs": 50, "description": 50,
    "images": 40, "local_images": 40, "url": 50,
}


def _get_lanczos():
    """兼容 Pillow 9 / 10 / 11 / 12 拿 LANCZOS resample 常量。"""
    if not HAS_PIL:
        return None
    # Pillow 9.1+ 推荐方式
    try:
        return PILImage.Resampling.LANCZOS
    except AttributeError:
        pass
    # Pillow 9.0 及以下
    val = getattr(PILImage, "LANCZOS", None)
    if val is not None:
        return val
    # 极端情况：返回整数常量 1 (LANCZOS) 仍能用
    return 1


def _make_excel_thumb_bytes(src_path: str) -> bytes | None:
    """从本地图生成 PNG 缩略图字节。失败返回 None 并打印原因。"""
    if not HAS_PIL:
        logger.warning(f"  Pillow 不可用，无法嵌入图片")
        return None
    if not src_path:
        return None
    p = Path(src_path)
    if not p.exists():
        logger.warning(f"  本地图片不存在：{src_path}")
        return None
    if p.stat().st_size < 100:
        logger.warning(f"  图片文件异常小：{src_path} ({p.stat().st_size}B)")
        return None
    try:
        with PILImage.open(src_path) as img:
            resample = _get_lanczos()
            try:
                if resample is not None:
                    img.thumbnail((THUMB_SIZE, THUMB_SIZE), resample)
                else:
                    img.thumbnail((THUMB_SIZE, THUMB_SIZE))
            except Exception:
                # 不传 resample 兜底
                img.thumbnail((THUMB_SIZE, THUMB_SIZE))

            if img.mode in ("RGBA", "LA", "P"):
                bg = PILImage.new("RGB", img.size, (255, 255, 255))
                if img.mode == "P":
                    img = img.convert("RGBA")
                mask = img.split()[-1] if img.mode in ("RGBA", "LA") else None
                bg.paste(img, mask=mask)
                img = bg
            elif img.mode != "RGB":
                img = img.convert("RGB")

            buf = BytesIO()
            img.save(buf, format="PNG", optimize=True)
            return buf.getvalue()
    except Exception as exc:
        logger.warning(f"  生成缩略图失败 {src_path}: {exc}")
        return None


class Storage:
    def __init__(self, output_cfg: dict):
        self.dir = Path(output_cfg.get("dir", "output"))
        self.dir.mkdir(parents=True, exist_ok=True)
        self.images_dir = Path(output_cfg.get("images_dir", self.dir / "images"))
        self.write_excel = output_cfg.get("excel", True)
        self.write_json = output_cfg.get("json", True)

    # ---------- 图片下载（带 Referer + 详细日志） ----------
    def download_images_for(self, products: Iterable[Product]) -> None:
        products = list(products)
        ok_count = 0
        empty_count = 0
        fail_samples: list[str] = []
        for p in tqdm(products, desc="下载图片"):
            try:
                if not p.images:
                    empty_count += 1
                    continue
                sub = self.images_dir / safe_filename(p.platform) / safe_filename(
                    p.bucket_type or "未分类"
                ) / safe_filename(p.product_id)
                local: list[str] = []
                for idx, url in enumerate(p.images):
                    try:
                        path = download_image(url, sub, prefix=f"{idx:02d}")
                        if path:
                            local.append(path)
                        elif len(fail_samples) < 5:
                            fail_samples.append(url)
                    except Exception as exc:
                        logger.warning(f"  下载图片异常 {url}: {exc}")
                p.local_images = local
                if local:
                    ok_count += 1
            except Exception as exc:
                logger.warning(f"  处理 {p.product_id} 图片异常：{exc}")
        logger.info(
            f"图片下载完成：{ok_count}/{len(products)} 件成功，"
            f"{empty_count} 件没有图片URL"
        )
        if fail_samples:
            logger.warning(f"  部分下载失败样例：")
            for u in fail_samples:
                logger.warning(f"    × {u}")

    # ---------- 主导出入口 ----------
    def save(self, products: list[Product]) -> dict[str, str]:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        results: dict[str, str] = {}

        if not products:
            logger.warning("没有可保存的商品。")
            return results

        # 概要
        n_with_local = sum(1 for p in products if p.local_images)
        logger.info(f"准备导出 {len(products)} 件商品（{n_with_local} 件有本地图）")

        if self.write_json:
            jpath = self.dir / f"products_{ts}.json"
            try:
                jpath.write_text(
                    json.dumps([p.to_dict() for p in products], ensure_ascii=False, indent=2),
                    encoding="utf-8",
                )
                results["json"] = str(jpath)
                logger.info(f"已写入 JSON: {jpath}")
            except Exception as exc:
                logger.error(f"写 JSON 失败 {jpath}: {exc}")

        if self.write_excel:
            xpath = self.dir / f"products_{ts}.xlsx"
            try:
                self._write_excel(xpath, products)
                results["excel"] = str(xpath)
                logger.info(f"已写入 Excel（含嵌入图片 + 中俄双语）: {xpath}")
            except PermissionError:
                alt = self.dir / f"products_{ts}_新.xlsx"
                logger.warning(f"Excel 文件被占用，改写到 {alt.name}")
                try:
                    self._write_excel(alt, products)
                    results["excel"] = str(alt)
                    logger.info(f"已写入 Excel: {alt}")
                except Exception as exc:
                    logger.error(f"写 Excel 失败 {alt}: {exc}")
            except Exception as exc:
                logger.error(f"写 Excel 失败 {xpath}: {exc}")

        return results

    # ---------- Excel ----------
    def _write_excel(self, path: Path, products: list[Product]) -> None:
        wb = Workbook()
        ws_all = wb.active
        ws_all.title = "ALL"
        self._fill_sheet(ws_all, products)

        by_type: dict[str, list[Product]] = defaultdict(list)
        for p in products:
            by_type[p.bucket_type or "未分类"].append(p)
        for type_name, items in by_type.items():
            sheet_name = safe_filename(type_name, max_len=28) or "未分类"
            sheet_name = self._unique_name(wb, sheet_name)
            ws = wb.create_sheet(sheet_name)
            self._fill_sheet(ws, items)

        by_bucket: dict[str, list[Product]] = defaultdict(list)
        for p in products:
            by_bucket[p.bucket_price or "未知"].append(p)
        for bucket, items in by_bucket.items():
            sheet_name = self._unique_name(wb, f"价_{safe_filename(bucket, max_len=24)}")
            ws = wb.create_sheet(sheet_name)
            self._fill_sheet(ws, items)

        wb.save(path)

    @staticmethod
    def _unique_name(wb: Workbook, base: str) -> str:
        name = base[:31] or "Sheet"
        existing = set(wb.sheetnames)
        if name not in existing:
            return name
        i = 2
        while f"{name[:28]}_{i}" in existing:
            i += 1
        return f"{name[:28]}_{i}"

    @staticmethod
    def _fill_sheet(ws, items: list[Product]) -> None:
        # 表头
        for col_idx, (_key, title, _fn) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "B2"

        for col_idx, (key, _title, _fn) in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(key, 16)

        embedded = 0
        no_local = 0
        thumb_failed = 0
        embed_failed = 0

        for row_idx, p in enumerate(items, start=2):
            ws.row_dimensions[row_idx].height = EXCEL_ROW_HEIGHT

            preview_path = p.local_images[0] if p.local_images else ""
            if not preview_path:
                no_local += 1
            else:
                thumb = _make_excel_thumb_bytes(preview_path)
                if not thumb:
                    thumb_failed += 1
                else:
                    try:
                        xlimg = XLImage(BytesIO(thumb))
                        xlimg.width = THUMB_SIZE
                        xlimg.height = THUMB_SIZE
                        ws.add_image(xlimg, f"A{row_idx}")
                        embedded += 1
                    except Exception as exc:
                        logger.warning(f"  嵌入图片失败 {p.product_id}: {exc}")
                        embed_failed += 1

            for col_idx, (key, _title, fn) in enumerate(COLUMNS, start=1):
                if key == PREVIEW_KEY:
                    continue
                try:
                    val = fn(p) if fn else ""
                except Exception:
                    val = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        if items:
            ws.auto_filter.ref = ws.dimensions

        # 详细诊断输出
        diag: list[str] = []
        if no_local:
            diag.append(f"无本地图 {no_local}")
        if thumb_failed:
            diag.append(f"缩略图失败 {thumb_failed}")
        if embed_failed:
            diag.append(f"嵌入失败 {embed_failed}")
        diag_str = (" | " + ", ".join(diag)) if diag else ""
        logger.info(f"  Sheet '{ws.title}'：嵌入 {embedded}/{len(items)} 张图{diag_str}")
